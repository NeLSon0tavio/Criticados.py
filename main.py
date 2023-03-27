import glob
import threading
import datetime
import imaplib
import email
import os
import zipfile
import openpyxl
import pandas as pd
from email.header import decode_header
import os

from dateutil.utils import today

resultadoC6 = 0
resultadopan = 0
resultadoSAFRA = 0
resultadoMERCANTIL = 0
resultadoDAYCOVAL = 0
resultadoFACTA = 0
resultado = 0
dia = ""
mes = ""
if datetime.date.today().day < 10:
    dia += "0"
dia += str(datetime.date.today().day)
if datetime.date.today().month < 10:
    mes += "0"
mes += str(datetime.date.today().month)
saida = os.getcwd() + "/Subir"
entrada = os.getcwd()


def baixarArquivos(quantEmails):
    imap = imaplib.IMAP4_SSL("imap.houseti.com.br")  # establish connection

    imap.login("nelson-santos@grupoamp.com.br", "!@#AMenorParcela2020")  # login

    status, messages = imap.select("INBOX")  # select inbox

    numOfMessages = int(messages[0])  # get number of messages

    for i in range(numOfMessages, numOfMessages - quantEmails, -1):
        res, msg = imap.fetch(str(i), "(RFC822)")
        for response in msg:
            if isinstance(response, tuple):
                msg = email.message_from_bytes(response[1])
                try:
                    subject, dateEmail = obtain_header(msg)
                except:
                    continue
                if msg.is_multipart():
                    for part in msg.walk():
                        body = ""
                        charset = part.get_content_charset()
                        if part.get_content_type() == "text/plain":
                            partStr = part.get_payload(decode=True)
                            body += partStr.decode(charset)
                        # content_type = part.get_content_type()
                        content_disposition = str(part.get("Content-Disposition"))
                        if "attachment" in content_disposition:
                            download_attachment(subject, part, dateEmail, i, numOfMessages)
    imap.close()
    # while resultadoC6 == 0:
    C6(dia, mes)
    # while resultadoSAFRA == 0:
    SAFRA(dia, mes)
    # while resultadoFACTA == 0:
    Facta(dia, mes)
    # while resultadoDAYCOVAL == 0:
    DAYCOVAL(dia, mes)
    # while resultadopan == 0:
    pan(dia, mes)
    # while resultadoMERCANTIL == 0:
    # MERCANTIL(dia, mes)


def clean(text):
    # clean text for creating a folder
    return "".join(c if c.isalnum() else "_" for c in text)


def obtain_header(msg):
    dateEmail = msg["Date"]
    subject, encoding = decode_header(msg["Subject"])[0]
    if isinstance(subject, bytes):
        subject = subject.decode(encoding)

    return subject, dateEmail


def download_attachment(subject, part, dateEmail, i, total):
    filename = part.get_filename()
    if filename:
        namefolder = ""
        if "SAFRA] - CRITICADO" in subject:
            namefolder = "SAFRA"
            if "NORMAL" in subject:
                namefolder += "/normal"
            else:
                namefolder += "/lewe"
        if "Relatorio de produção de correspondentes Completo" in subject:
            namefolder = "Pan"
        if "MERCANTIL] CRITICADOS" in subject:
            namefolder = "Mercantil"
        if "FACTA" in filename:
            namefolder = "FACTA"
        if "criticados cartao" in filename:
            namefolder = "DAYCOVAL"
        if "criticados.csv" == filename and "C6" in subject:
            namefolder = "C6"
            if "JSS" in subject:
                namefolder += "/JSS"
            if "VIEIRA" in subject:
                namefolder += "/VIEIRA"
            if "SILVASEG" in subject:
                namefolder += "/SILVASEG"
            if "BRACEIRO" in subject:
                namefolder += "/BRACEIRO"
            if "MAUA DOIS" in subject:
                namefolder += "/MAUA DOIS"
            if "JR" in subject:
                namefolder += "/JR"
            if "DSA" in subject:
                namefolder += "/DSA"
            if "AGUIAR" in subject:
                namefolder += "/AGUIAR"
            if "FHBS" in subject:
                namefolder += "/FHBS"
        if "FACTA" in filename:
            namefolder = "FACTA"
            if "6000041" in subject:
                namefolder += "/6000041"
            if "600061" in subject:
                namefolder += "/600061"
            if "930010" in subject:
                namefolder += "/930010"
            if "600060" in subject:
                namefolder += "/600060"
            if "600062" in subject:
                namefolder += "/600062"
            if "600059" in subject:
                namefolder += "/600059"
            if "92834" in subject:
                namefolder += "/92834"
            if "53599" in subject:
                namefolder += "/53599"
            if "54538" in subject:
                namefolder += "/54538"
            if "600053" in subject:
                namefolder += "/600053"
        folder_name = clean(subject)
        if namefolder != "":
            time = str(dateEmail[4:22]).strip().replace(":", "-")
            if "Pan" in namefolder:
                filename = filename[:47] + time + ".zip"
            else:
                filename = filename.replace(".csv", " ") + time + ".csv"
            if not os.path.isdir(folder_name):
                filepath = os.path.join(namefolder, filename)
                try:
                    open(filepath, "wb").write(part.get_payload(decode=True))
                except FileNotFoundError:
                    try:
                        os.mkdir(os.path.join(namefolder))
                        open(filepath, "wb").write(part.get_payload(decode=True))
                    except FileNotFoundError:
                        os.mkdir(namefolder.split("/")[0])
                        os.mkdir(os.path.join(namefolder))
                        open(filepath, "wb").write(part.get_payload(decode=True))
                print("Baixando arquivo N° ", (total - i), " referente a financeira -> ", namefolder)


# data = {'Promotora': row1,
#         'FinId': row2,
#         'CcNumeroContrato': row3,
#         'CcCliCpf': row4,
#         'CcCliNome': row5,
#         'CcNomeDigitadorBanco': row6,
#         'CcDataStatus': row7,
#         'CcStatusBanco': row8,
#         'CcDataCadastro': row9}
# df = pd.DataFrame(data, columns=['Promotora', 'FinId', 'CcNumeroContrato',
#                                  'CcCliCpf', 'CcCliNome', 'CcNomeDigitadorBanco',
#                                  'CcDataStatus',
#                                  'CcStatusBanco', 'CcDataCadastro'])

# try:
#     df.to_csv(saida + f'/Subir SAFRA {dia}{mes}.csv', index=False, sep=";")
# except OSError:
#     os.mkdir(os.path.join("Subir"))
#     df.to_csv(saida + f'/Subir SAFRA {dia}{mes}.csv', index=False, sep=";")

# else:
# content_type = msg.get_content_type()
# body = msg.get_payload(decode=True).decode()

def C6(dia, mes):
    caminho = entrada + '/C6'
    files = []
    files_braceiro = glob.glob(caminho + "/BRACEIRO" + "/*.csv")
    files_dsa = glob.glob(caminho + "/DSA" + "/*.csv")
    files_fhbs = glob.glob(caminho + "/FHBS" + "/*.csv")
    files_jr = glob.glob(caminho + "/JR" + "/*.csv")
    files_jss = glob.glob(caminho + "/JSS" + "/*.csv")
    files_maua_dois = glob.glob(caminho + "/MAUA DOIS" + "/*.csv")
    files_silvaseg = glob.glob(caminho + "/SILVASEG" + "/*.csv")
    files_vieira = glob.glob(caminho + "/VIEIRA" + "/*.csv")
    files.append(glob.glob(caminho + "/AGUIAR" + "/*.csv"))
    files.append(files_braceiro)
    files.append(files_dsa)
    files.append(files_fhbs)
    files.append(files_jr)
    files.append(files_jss)
    files.append(files_maua_dois)
    files.append(files_silvaseg)
    files.append(files_vieira)
    months = ["jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]
    selected = []

    row1 = []
    row2 = []
    row3 = []
    row4 = []
    row5 = []
    row6 = []
    row7 = []
    row8 = []
    row9 = []

    for file in files:
        for filename in file:
            if f"{dia} {months[datetime.date.today().month - 1].title()}" in filename:  # 9821
                caminho2 = filename.replace("\\", "/")
                file2 = pd.read_csv(caminho2, encoding='latin1', on_bad_lines='skip', skiprows=0, sep=';')

                for linha in range(file2[str(file2.keys()[1])].size):  # mudar para coluna
                    if str(file2[str(file2.keys()[1])][linha]) == "nan":
                        break
                    gNumber = file2[str(file2.keys()[6])][linha]
                    if gNumber != "None" and str(gNumber) != "":
                        if "cancelada" not in str(gNumber).lower() and "reprovada" not in str(gNumber).lower() and \
                                "andamento - cadastro de proposta" not in str(gNumber).lower():
                            add = True
                            for i in range(len(selected)):
                                if str(file2[str(file2.keys()[0])][linha]) == str(selected[i][0]):
                                    add = False
                            if add:
                                row_selected = [str(file2[str(file2.keys()[0])][linha]),
                                                str(file2[str(file2.keys()[1])][linha]),
                                                str(file2[str(file2.keys()[2])][linha]),
                                                str(file2[str(file2.keys()[3])][linha]),
                                                str(file2[str(file2.keys()[4])][linha]),
                                                str(file2[str(file2.keys()[5])][linha]),
                                                str(file2[str(file2.keys()[6])][linha]),
                                                str(file2[str(file2.keys()[7])][linha]), "C6 consig",
                                                " ",
                                                str(file2[str(file2.keys()[8])][linha])]

                                selected.append(row_selected)
    bookfinal = openpyxl.Workbook()
    bookfinal.iso_dates = True
    sheetfinal = bookfinal.active

    for x in range(0, len(selected)):
        for coluna in sheetfinal["A" + str(x + 1) + ":K" + str(x + 1)]:
            if "data status" not in selected[x][5]:
                coluna[0].value = selected[x][0]
                year = ""
                if not selected[x][5].split("/")[1].isdigit():
                    mes = ""
                    if months.index(selected[x][5].split("/")[1]) + 1 < 10:
                        mes += "0"
                    mes += str(months.index(selected[x][5].split("/")[1]) + 1)
                    year += str(today().year) + "-" + mes + "-" + str(selected[x][5].split("/")[0])
                else:
                    year += str(today().year) + "-" + selected[x][5].split("/")[1] + "-" + str(
                        selected[x][5].split("/")[0])
                year.replace("/", "-")
                a = ""
                if selected[x][7] == "" or selected[x][7] == '-':
                    a = selected[x][4]
                else:
                    a = selected[x][7]
                coluna[0].value = a
                coluna[1].value = "29"
                coluna[2].value = selected[x][0]
                coluna[3].value = selected[x][1]
                coluna[4].value = selected[x][2]
                coluna[5].value = selected[x][3]
                coluna[6].value = year
                coluna[7].value = selected[x][6]
                coluna[8].value = year
                row1.append(a)
                row2.append("29")
                row3.append(str(selected[x][0]))
                row4.append(selected[x][1])
                row5.append(selected[x][2])
                row6.append(selected[x][7])
                row7.append(year)
                row8.append(selected[x][6])
                row9.append(year)
    if len(row3) != 0:
        data = {'Promotora': row1,
                'FinId': row2,
                'CcNumeroContrato': row3,
                'CcCliCpf': row4,
                'CcCliNome': row5,
                'CcNomeDigitadorBanco': row6,
                'CcDataStatus': row7,
                'CcStatusBanco': row8,
                'CcDataCadastro': row9}
        df = pd.DataFrame(data, columns=['Promotora', 'FinId', 'CcNumeroContrato',
                                         'CcCliCpf', 'CcCliNome', 'CcNomeDigitadorBanco', 'CcDataStatus',
                                         'CcStatusBanco', 'CcDataCadastro'])

        try:
            df.to_csv(saida + f'/Subir C6 {dia}{mes}.csv', index=False, sep=";", encoding="latin1")
        except OSError:
            os.mkdir(os.path.join("Subir"))
            df.to_csv(saida + f'/Subir C6 {dia}{mes}.csv', index=False, sep=";", encoding="latin1")
        # print(f"Terminando C6 com {len(row3)} contratos")
    resultadoC6 = len(row3)


# def MERCANTIL(dia, mes):
#     caminho = entrada + '/MERCANTIL'
#     files = [glob.glob(caminho + "/*.csv")]
#     months = ["jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]
#     selected = []
#
#     row1 = []
#     row2 = []
#     row3 = []
#     row4 = []
#     row5 = []
#     row6 = []
#     row7 = []
#     row8 = []
#     row9 = []
#
#     for file in files:
#         for filename in file:
#             if f"{dia} {mes}" in filename:  # 9821
#                 caminho2 = filename.replace("\\", "/")
#                 file2 = pd.read_csv(caminho2, encoding='latin-1', on_bad_lines='skip', skiprows=0, sep=';')
#                 for linha in range(file2[str(file2.keys()[1])].size):  # mudar para coluna
#                     if str(file2[str(file2.keys()[1])][linha]) == "nan":
#                         break
#                     gNumber = file2[str(file2.keys()[6])][linha]
#                     if gNumber != "None" and str(gNumber) != "":
#                         if "cancelada" not in str(gNumber).lower() and "reprovada" not in str(gNumber).lower() and \
#                                 "andamento - cadastro de proposta" not in str(gNumber).lower():
#                             add = True
#                             for i in range(len(selected)):
#                                 if str(file2[str(file2.keys()[0])][linha]) == str(selected[i][0]):
#                                     add = False
#                             if add:
#                                 row_selected = [str(file2[str(file2.keys()[0])][linha]),
#                                                 str(file2[str(file2.keys()[1])][linha]),
#                                                 str(file2[str(file2.keys()[2])][linha]),
#                                                 str(file2[str(file2.keys()[3])][linha]),
#                                                 str(file2[str(file2.keys()[4])][linha]),
#                                                 str(file2[str(file2.keys()[5])][linha]),
#                                                 str(file2[str(file2.keys()[6])][linha]),
#                                                 str(file2[str(file2.keys()[7])][linha]), "MERCANTIL",
#                                                 " ",
#                                                 str(file2[str(file2.keys()[8])][linha])]
#
#                                 selected.append(row_selected)
#     bookfinal = openpyxl.Workbook()
#     bookfinal.iso_dates = True
#     sheetfinal = bookfinal.active
#
#     for x in range(0, len(selected)):
#         for coluna in sheetfinal["A" + str(x + 1) + ":K" + str(x + 1)]:
#             if "data status" not in selected[x][5]:
#                 coluna[0].value = selected[x][0]
#                 year = ""
#                 if not selected[x][5].split("/")[1].isdigit():
#                     mes = ""
#                     if months.index(selected[x][5].split("/")[1]) + 1 < 10:
#                         mes += "0"
#                     mes += str(months.index(selected[x][5].split("/")[1]) + 1)
#                     year += str(today().year) + "-" + mes + "-" + str(selected[x][5].split("/")[0])
#                 else:
#                     year += str(today().year) + "-" + selected[x][5].split("/")[1] + "-" + str(
#                         selected[x][5].split("/")[0])
#                 year.replace("/", "-")
#                 a = ""
#                 if selected[x][7] == "" or selected[x][7] == '-':
#                     a = selected[x][4]
#                 else:
#                     a = selected[x][7]
#                 row1.append(a)
#                 row2.append("16")
#                 row3.append(str(selected[x][0]))
#                 row4.append(selected[x][1])
#                 row5.append(selected[x][2])
#                 row6.append(selected[x][3])
#                 row7.append(year)
#                 row8.append(selected[x][6])
#                 row9.append(year)
#     if len(row3) != 0:
#
#         data = {'Promotora': row1,
#                 'FinId': row2,
#                 'CcNumeroContrato': row3,
#                 'CcCliCpf': row4,
#                 'CcCliNome': row5,
#                 'CcNomeDigitadorBanco': row6,
#                 'CcDataStatus': row7,
#                 'CcStatusBanco': row8,
#                 'CcDataCadastro': row9}
#         print(row1)
#         df = pd.DataFrame(data, columns=['Promotora', 'FinId', 'CcNumeroContrato',
#                                          'CcCliCpf', 'CcCliNome', 'CcNomeDigitadorBanco', 'CcDataStatus',
#                                          'CcStatusBanco', 'CcDataCadastro'])
#
#         try:
#             df.to_csv(saida + f'/Subir MERCANTIL {dia}{mes}.csv', index=False, sep=";")
#         except OSError:
#             os.mkdir(os.path.join("Subir"))
#             df.to_csv(saida + f'/Subir MERCANTIL {dia}{mes}.csv', index=False, sep=";")
#     print(f"Terminando MERCANTIL com {len(row3)} contratos")
#     return len(row3)


def Facta(dia, mes):
    caminho = entrada + r"/Facta"
    files = []
    files_53599 = glob.glob(caminho + "/53599" + "/*.csv")
    files_54538 = glob.glob(caminho + "/54538" + "/*.csv")
    files_92834 = glob.glob(caminho + "/92834" + "/*.csv")
    files_600053 = glob.glob(caminho + "/600053" + "/*.csv")
    files_600059 = glob.glob(caminho + "/600059" + "/*.csv")
    files_600060 = glob.glob(caminho + "/600060" + "/*.csv")
    files_600061 = glob.glob(caminho + "/600061" + "/*.csv")
    files_600062 = glob.glob(caminho + "/600062" + "/*.csv")
    files_930010 = glob.glob(caminho + "/930010" + "/*.csv")
    files_6000041 = glob.glob(caminho + "/6000041" + "/*.csv")
    files.append(files_53599)
    files.append(files_54538)
    files.append(files_92834)
    files.append(files_600053)
    files.append(files_600059)
    files.append(files_600060)
    files.append(files_600061)
    files.append(files_600062)
    files.append(files_930010)
    files.append(files_6000041)
    months = ["jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]
    selected = []

    row1 = []
    row2 = []
    row3 = []
    row4 = []
    row5 = []
    row6 = []
    row7 = []
    row8 = []
    row9 = []

    for file in files:
        for filename in file:
            if f"{dia} {months[datetime.date.today().month - 1].title()}" in filename:  # 9821
                caminho2 = filename.replace("\\", "/")
                file2 = pd.read_csv(caminho2, encoding='latin-1', on_bad_lines='skip', skiprows=0, sep=';')
                for linha in range(file2[str(file2.keys()[1])].size):  # mudar para coluna
                    if str(file2[str(file2.keys()[1])][linha]) == "nan":
                        break
                    gNumber = file2[str(file2.keys()[5])][linha]
                    if gNumber != "None":
                        if "rep" not in str(gNumber).lower() and \
                                "can" not in str(gNumber).lower():
                            add = True
                            for i in range(len(selected)):
                                if str(file2[str(file2.keys()[0])][linha]) == str(selected[i][0]):
                                    add = False
                            if add:
                                row_selected = [str(file2[str(file2.keys()[0])][linha]),
                                                str(file2[str(file2.keys()[1])][linha]),
                                                str(file2[str(file2.keys()[2])][linha]),
                                                str(file2[str(file2.keys()[3])][linha]),
                                                str(file2[str(file2.keys()[4])][linha]),
                                                str(file2[str(file2.keys()[5])][linha]),
                                                str(file2[str(file2.keys()[6])][linha]),
                                                str(file2[str(file2.keys()[7])][linha]),
                                                "FACTA"
                                                " ",
                                                str(file2[str(file2.keys()[7])][linha])]
                                selected.append(row_selected)

    print(selected[0])
    for x in range(0, len(selected)):
        row1.append(str(selected[x][6]))
        row2.append("21")
        row3.append(str(selected[x][0]))
        row4.append(selected[x][1])
        row5.append(selected[x][2])
        row6.append(selected[x][3])
        row7.append(selected[x][4])
        row8.append(selected[x][5])
        row9.append(selected[x][9])
    if len(row3) != 0:
        data = {'Promotora': row1,
                'FinId': row2,
                'CcNumeroContrato': row3,
                'CcCliCpf': row4,
                'CcCliNome': row5,
                'CcNomeDigitadorBanco': row6,
                'CcDataStatus': row7,
                'CcStatusBanco': row8,
                'CcDataCadastro': row9}
        df = pd.DataFrame(data, columns=['Promotora', 'FinId', 'CcNumeroContrato',
                                         'CcCliCpf', 'CcCliNome', 'CcNomeDigitadorBanco', 'CcDataStatus',
                                         'CcStatusBanco', 'CcDataCadastro'])
        df.to_csv(saida + f'/Subir Facta {dia}{mes}.csv', index=False, sep=";", encoding="latin-1")
    print(f"Terminando FACTA com {len(row3)} contratos")


# return len(row3)


def DAYCOVAL(dia, mes):
    caminho = entrada + r"\DAYCOVAL"
    caminho = caminho.replace('\\', '/')
    months = ["jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]
    files = glob.glob(caminho + "/*.csv")

    selected = []

    row1 = []
    row2 = []
    row3 = []
    row4 = []
    row5 = []
    row6 = []
    row7 = []
    row8 = []
    row9 = []

    for filename in files:
        if f"{dia} {months[datetime.date.today().month - 1].title()}" in filename:  # 9821
            caminho2 = filename.replace("\\", "/")
            file2 = pd.read_csv(caminho2, encoding='Utf-8', on_bad_lines='skip', skiprows=0, sep=';')
            for linha in range(file2[str(file2.keys()[1])].size):  # mudar para coluna
                if str(file2[str(file2.keys()[1])][linha]) == "nan":
                    break
                gNumber = file2[str(file2.keys()[6])][linha]
                if gNumber != "None":
                    if "rep" not in str(gNumber).lower() and \
                            "can" not in str(gNumber).lower() and \
                            "rec" not in str(gNumber).lower():
                        add = True
                        for i in range(len(selected)):
                            if str(file2[str(file2.keys()[0])][linha]) == str(selected[i][0]):
                                add = False
                        if add:
                            row_selected = [str(file2[str(file2.keys()[0])][linha]),
                                            str(file2[str(file2.keys()[1])][linha]),
                                            str(file2[str(file2.keys()[2])][linha]),
                                            str(file2[str(file2.keys()[3])][linha]).replace("nan", ""),
                                            str(file2[str(file2.keys()[4])][linha]).replace("nan", ""),
                                            str(file2[str(file2.keys()[5])][linha]),
                                            str(file2[str(file2.keys()[6])][linha]),
                                            str(file2[str(file2.keys()[7])][linha]), "Daycoval",
                                            "",
                                            str(file2[str(file2.keys()[10])][linha])]

                            selected.append(row_selected)

    for x in range(0, len(selected)):
        year = ""
        if not selected[x][5].split("/")[1].isdigit():
            mes = ""
            if months.index(selected[x][5].split("/")[1]) + 1 < 10:
                mes += "0"
            mes += str(months.index(selected[x][5].split("/")[1]) + 1)
            year += str(today().year) + "-" + mes + "-" + str(selected[x][5].split("/")[0])
        else:
            year += str(today().year) + "-" + selected[x][5].split("/")[1] + "-" + str(
                selected[x][5].split("/")[0])
        year.replace("/", "-")
        row1.append(selected[x][7])
        row2.append("8")
        row3.append(selected[x][0])
        row4.append(selected[x][1])
        row5.append(selected[x][2])
        row6.append(selected[x][3])
        row7.append(year)
        row8.append(selected[x][6])
        row9.append(year)
    if len(row3) != 0:
        data = {'Promotora': row1,
                'FinId': row2,
                'CcNumeroContrato': row3,
                'CcCliCpf': row4,
                'CcCliNome': row5,
                'CcNomeDigitadorBanco': row6,
                'CcDataStatus': row7,
                'CcStatusBanco': row8,
                'CcDataCadastro': row9}
        df = pd.DataFrame(data, columns=['Promotora', 'FinId', 'CcNumeroContrato',
                                         'CcCliCpf', 'CcCliNome', 'CcNomeDigitadorBanco', 'CcDataStatus',
                                         'CcStatusBanco', 'CcDataCadastro'])
        df.to_csv(saida + f'/Subir Daycoval {dia}{mes}.csv', index=False, sep=";")
        print(f"Terminando DAYCOVAL com {len(row3)} contratos")


def SAFRA(dia, mes):
    files = []
    files_normal = glob.glob(entrada + "/SAFRA/NORMAL" + "/*.csv")
    files.append(files_normal)
    files_lewe = glob.glob(entrada + "/SAFRA/LEWE" + "/*.csv")
    files.append(files_lewe)
    months = ["jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]
    selected = []

    row1 = []
    row2 = []
    row3 = []
    row4 = []
    row5 = []
    row6 = []
    row7 = []
    row8 = []
    row9 = []

    for file in files:
        for filename in file:
            # if f"{dia} {mes}" in filename:  # 9821
            caminho2 = filename.replace("\\", "/")
            file2 = pd.read_csv(caminho2, encoding='latin-1', on_bad_lines='skip', skiprows=0, sep=';')
            for linha in range(file2[str(file2.keys()[1])].size):  # mudar para coluna
                if str(file2[str(file2.keys()[1])][linha]) == "nan":
                    break
                gNumber = file2[str(file2.keys()[7])][linha]
                # print(gNumber)
                if str(gNumber) != "None" and gNumber != "+++" and gNumber != "0":
                    if "recusada" not in str(gNumber).lower() and "cancelado" not in str(gNumber).lower() and \
                            "expirado" not in str(gNumber).lower():
                        row_selected = [str(file2[str(file2.keys()[0])][linha]),
                                        str(file2[str(file2.keys()[1])][linha]),
                                        str(file2[str(file2.keys()[2])][linha]),
                                        str(file2[str(file2.keys()[3])][linha]),
                                        str(file2[str(file2.keys()[4])][linha]),
                                        str(file2[str(file2.keys()[5])][linha]),
                                        str(file2[str(file2.keys()[6])][linha]),
                                        str(file2[str(file2.keys()[7])][linha]), "SAFRA",
                                        " ",
                                        str(file2[str(file2.keys()[8])][linha])]
                        selected.append(row_selected)
    bookfinal = openpyxl.Workbook()
    bookfinal.iso_dates = True
    sheetfinal = bookfinal.active

    for x in range(0, len(selected)):
        for coluna in sheetfinal["A" + str(x + 1) + ":K" + str(x + 1)]:
            year = ""
            if not selected[x][6].split("/")[1].isdigit():
                mes = ""
                if months.index(selected[x][6].split("/")[1]) + 1 < 10:
                    mes += "0"
                mes += str(months.index(selected[x][6].split("/")[1]) + 1)
                year += str(today().year) + "-" + mes + "-" + str(selected[x][6].split("/")[0])
            else:
                year += str(today().year) + "-" + selected[x][6].split("/")[1] + "-" + str(
                    selected[x][6].split("/")[0])
            year.replace("/", "-")
            # a = ""
            # if selected[x][7] == "" or selected[x][7] == '-':
            #     a = selected[x][4]
            # else:
            #     a = selected[x][7]
            row1.append(str(selected[x][0]))
            row2.append("2")
            row3.append(str(selected[x][2]))
            row4.append(selected[x][3])
            row5.append(selected[x][4])
            row6.append(selected[x][5])
            row7.append(year)
            row8.append(selected[x][7])
            row9.append(year)
    if len(row3) != 0:
        data = {'Promotora': row1,
                'FinId': row2,
                'CcNumeroContrato': row3,
                'CcCliCpf': row4,
                'CcCliNome': row5,
                'CcNomeDigitadorBanco': row6,
                'CcDataStatus': row7,
                'CcStatusBanco': row8,
                'CcDataCadastro': row9}
        df = pd.DataFrame(data, columns=['Promotora', 'FinId', 'CcNumeroContrato',
                                         'CcCliCpf', 'CcCliNome', 'CcNomeDigitadorBanco', 'CcDataStatus',
                                         'CcStatusBanco', 'CcDataCadastro'])

        df.to_csv(saida + f'/Subir SAFRA {dia}{mes}.csv', index=False, sep=";", enconding="latin1")
    print(f"Terminando SAFRA com {len(row3)} contratos")
    return len(row3)


def pan(dia, mes):
    diretorio = entrada + "/Pan"
    for arquivo in os.listdir(diretorio):
        if arquivo.endswith('.zip'):
            try:
                caminho_completo = os.path.join(diretorio, arquivo)
                with zipfile.ZipFile(caminho_completo, 'r') as zip_ref:
                    zip_ref.extractall(diretorio)
                nome_arquivo_extraido = f"RelatorioProducaoCorrespondentesCompletoDIRECTX.CSV"
                novo_nome_arquivo_extraido = 'RelatorioProducaoCorrespondentesCompletoDIRECTX' + \
                                             caminho_completo.replace(" ", "$").split("$")[
                                                 3] + " " + dia + " " + mes + ".CSV"
                caminho_completo_antigo = os.path.join(diretorio, nome_arquivo_extraido)
                caminho_completo_novo = os.path.join(diretorio, novo_nome_arquivo_extraido)
                os.rename(caminho_completo_antigo, caminho_completo_novo)
                os.remove(caminho_completo)
            except FileExistsError:
                pass

    caminho = entrada
    files = []
    files_lewe = glob.glob(caminho + "/Pan" + "/*.csv")
    files.append(files_lewe)
    months = ["jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]
    selected = []

    row1 = []
    row2 = []
    row3 = []
    row4 = []
    row5 = []
    row6 = []
    row7 = []
    row8 = []
    row9 = []
    for file in files:
        for filename in file:
            if f"{dia} {mes}" in filename:  # 9821
                caminho2 = filename.replace("\\", "/")
                file2 = pd.read_csv(caminho2, encoding='latin-1', on_bad_lines='skip', skiprows=0, sep=';')
                for linha in range(file2[str(file2.keys()[1])].size):  # mudar para coluna
                    if str(file2[str(file2.keys()[1])][linha]) == "nan":
                        break
                    gNumber = file2[str(file2.keys()[7])][linha] + " - " + file2[str(file2.keys()[9])][
                        linha]  # pegar somente o primeiro campo e fazer o filtro com
                    if str(gNumber) != "None" and gNumber != " " and gNumber != "0":
                        if "reprovada" not in str(gNumber).lower() and "cancelad" not in str(gNumber).lower() and \
                                "expirado" not in str(gNumber).lower():
                            print(str(file2.keys()[16]))
                            row_selected = [str(file2[str(file2.keys()[0])][linha]),
                                            str(file2[str(file2.keys()[1])][linha]),
                                            str(file2[str(file2.keys()[20])][linha]),
                                            str(file2[str(file2.keys()[22])][linha]),
                                            str(str(file2[str(file2.keys()[20])][linha])).replace(".", "").replace("-",
                                                                                                                   "") + "_00" + str(
                                                file2[str(file2.keys()[16])][linha]),
                                            str(file2[str(file2.keys()[5])][linha]),
                                            str(file2[str(file2.keys()[6])][linha]),
                                            str(file2[str(file2.keys()[7])][linha]), "Pan",
                                            " ",
                                            str(file2[str(file2.keys()[8])][linha])]
                            selected.append(row_selected)
    bookfinal = openpyxl.Workbook()
    bookfinal.iso_dates = True
    sheetfinal = bookfinal.active

    for x in range(0, len(selected)):
        for coluna in sheetfinal["A" + str(x + 1) + ":K" + str(x + 1)]:
            year = ""
            if not selected[x][6].split("/")[1].isdigit():
                mes = ""
                if months.index(selected[x][6].split("/")[1]) + 1 < 10:
                    mes += "0"
                mes += str(months.index(selected[x][6].split("/")[1]) + 1)
                year += str(today().year) + "-" + mes + "-" + str(selected[x][6].split("/")[0])
            else:
                year += str(today().year) + "-" + selected[x][6].split("/")[1] + "-" + str(selected[x][6].split("/")[0])
            year.replace("/", "-")
            # a = ""
            # if selected[x][7] == "" or selected[x][7] == '-':
            #     a = selected[x][4]
            # else:
            #     a = selected[x][7]
            row1.append(str(selected[x][0]))
            row2.append("2")
            row3.append(str(selected[x][1]))
            row4.append(selected[x][2])
            row5.append(selected[x][3])
            row6.append(selected[x][4])
            row7.append(year)
            row8.append(selected[x][6])
            row9.append(year)
    if len(row3) != 0:
        data = {'Promotora': row1,
                'FinId': row2,
                'CcNumeroContrato': row3,
                'CcCliCpf': row4,
                'CcCliNome': row5,
                'CcNomeDigitadorBanco': row6,
                'CcDataStatus': row7,
                'CcStatusBanco': row8,
                'CcDataCadastro': row9}
        df = pd.DataFrame(data, columns=['Promotora', 'FinId', 'CcNumeroContrato',
                                         'CcCliCpf', 'CcCliNome', 'CcNomeDigitadorBanco', 'CcDataStatus',
                                         'CcStatusBanco', 'CcDataCadastro'])
        df.to_csv(saida + f'/Subir Pan {dia}{mes}.csv', index=False, sep=";", encoding="latin1")
        print(f"Terminando PAN com {len(row3)} contratos")
        return len(row3)


#
baixarArquivos(1000)
# C6(dia, mes)
# Facta(dia, mes)
# SAFRA(dia, mes)
# DAYCOVAL(dia, mes)
# pan(dia, mes)

# MERCANTIL(dia,mes)
